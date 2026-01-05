#!/usr/bin/env python3
"""
Excel to SQL INSERT statements
No API needed - pure SQL output
"""

import openpyxl
import os

# ============================================
# CONFIGURATION
# ============================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, "derskonular.xlsx")
OUTPUT_SQL = os.path.join(SCRIPT_DIR, "update_topics_year_data.sql")

# Subject mapping
SUBJECT_MAPPING = {
    'Matematik': 'MAT',
    'Fizik': 'FIZ',
    'Kimya': 'KIM',
    'Biyoloji': 'BIO',
    'Coğrafya': 'COG',
    'Tarih': 'TAR',
    'Türkçe': 'TUR',
    'Edebiyat': 'EDB',
    'Felsefe': 'FEL',
    'Din': 'DIN',
    'Mantık': 'MAN',
    'Psikoloji': 'PSI',
    'Sosyoloji': 'SOS'
}

def clean_value(val):
    """Clean value for SQL"""
    if val is None or val == '':
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def escape_sql(text):
    """Escape single quotes for SQL"""
    if not text:
        return ''
    return str(text).replace("'", "''")

def main():
    print("=" * 70)
    print("📊 EXCEL TO SQL GENERATOR")
    print("=" * 70)
    
    # Load Excel
    print("\n📁 Loading Excel...")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb['Sayfa1']
        print(f"✅ Loaded: {ws.max_row - 3} rows")
    except Exception as e:
        print(f"❌ Failed: {e}")
        return
    
    # Generate SQL
    print("\n📝 Generating SQL statements...")
    
    sql_statements = []
    sql_statements.append("-- ============================================")
    sql_statements.append("-- UPDATE TOPICS WITH YEAR DATA")
    sql_statements.append("-- Generated from Excel: derskonular.xlsx")
    sql_statements.append("-- ============================================")
    sql_statements.append("")
    sql_statements.append("BEGIN;")
    sql_statements.append("")
    
    processed = 0
    skipped = 0
    
    for row_idx in range(4, ws.max_row + 1):
        # Read row
        konu_id = ws.cell(row_idx, 1).value
        sinif = ws.cell(row_idx, 2).value
        ders = ws.cell(row_idx, 3).value
        ana_konu = ws.cell(row_idx, 4).value
        
        if not all([konu_id, sinif, ders, ana_konu]):
            skipped += 1
            continue
        
        # Map subject
        subject_code = SUBJECT_MAPPING.get(ders)
        if not subject_code:
            skipped += 1
            continue
        
        # Parse year data
        years = [2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025]
        year_values = {}
        
        for i, year in enumerate(years):
            col_idx = 5 + i
            q_count = clean_value(ws.cell(row_idx, col_idx).value)
            year_values[f'q_{year}'] = q_count
        
        # Create topic code (matching pattern)
        topic_code = f"{subject_code}_{konu_id}"
        
        # Escape name
        safe_name = escape_sql(ana_konu)
        
        # Generate UPDATE statement
        sql = f"""UPDATE topics SET
    q_2018 = {year_values['q_2018']},
    q_2019 = {year_values['q_2019']},
    q_2020 = {year_values['q_2020']},
    q_2021 = {year_values['q_2021']},
    q_2022 = {year_values['q_2022']},
    q_2023 = {year_values['q_2023']},
    q_2024 = {year_values['q_2024']},
    q_2025 = {year_values['q_2025']}
WHERE code = '{topic_code}';
"""
        sql_statements.append(sql)
        processed += 1
        
        # Progress
        if row_idx % 100 == 0:
            print(f"   Processed {row_idx - 3}/{ws.max_row - 3}...")
    
    # Add frequency calculation
    sql_statements.append("")
    sql_statements.append("-- ============================================")
    sql_statements.append("-- UPDATE EXAM FREQUENCY FOR ALL TOPICS")
    sql_statements.append("-- ============================================")
    sql_statements.append("")
    sql_statements.append("-- Calculate frequency (may take 30 seconds)")
    sql_statements.append("SELECT update_all_exam_frequencies(5);")
    sql_statements.append("")
    sql_statements.append("COMMIT;")
    sql_statements.append("")
    sql_statements.append("-- ============================================")
    sql_statements.append("-- VERIFICATION")
    sql_statements.append("-- ============================================")
    sql_statements.append("")
    sql_statements.append("SELECT COUNT(*) as total_topics,")
    sql_statements.append("       COUNT(*) FILTER (WHERE q_2024 > 0) as with_2024_data,")
    sql_statements.append("       COUNT(*) FILTER (WHERE q_2023 > 0) as with_2023_data,")
    sql_statements.append("       COUNT(*) FILTER (WHERE exam_frequency IS NOT NULL) as with_frequency")
    sql_statements.append("FROM topics;")
    sql_statements.append("")
    sql_statements.append("-- Sample topics with frequency")
    sql_statements.append("SELECT name_tr, q_2023, q_2024, exam_frequency")
    sql_statements.append("FROM topics")
    sql_statements.append("WHERE exam_frequency IS NOT NULL")
    sql_statements.append("ORDER BY exam_frequency DESC")
    sql_statements.append("LIMIT 10;")
    
    # Write to file
    with open(OUTPUT_SQL, 'w', encoding='utf-8') as f:
        f.write('\n'.join(sql_statements))
    
    print(f"\n✅ Generated {processed} UPDATE statements")
    print(f"⚠️  Skipped {skipped} rows")
    print(f"\n📄 SQL file created: {OUTPUT_SQL}")
    print(f"   Size: {os.path.getsize(OUTPUT_SQL) / 1024:.1f} KB")
    
    print("\n" + "=" * 70)
    print("✅ SQL GENERATION COMPLETE!")
    print("=" * 70)
    print("\nNext steps:")
    print("1. Open Supabase SQL Editor")
    print("2. Copy/paste the SQL file content")
    print("3. Run the entire script")
    print("4. Check verification results")

if __name__ == "__main__":
    main()