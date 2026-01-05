#!/usr/bin/env python3
"""
Excel to SQL with Smart Name Matching
Fuzzy matching between Excel and Database topic names
"""

import openpyxl
import os
import re

# ============================================
# CONFIGURATION
# ============================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, "derskonular.xlsx")
OUTPUT_SQL = os.path.join(SCRIPT_DIR, "update_topics_smart_match.sql")

# Subject mapping
SUBJECT_MAPPING = {
    'Matematik': 'MAT',
    'Fizik': 'FIZ',
    'Kimya': 'KIM',
    'Biyoloji': 'BIO',
    'Coğrafya': 'COG',
    'Tarih': 'TAR',
    'Türkçe': 'TUR',
    'Edebiyat': 'EDB',
    'Felsefe': 'FEL',
    'Din': 'DIN',
    'Mantık': 'MAN',
    'Psikoloji': 'PSI',
    'Sosyoloji': 'SOS'
}

# Grade mapping
GRADE_MAPPING = {
    '9. Sınıf': '9',
    '10. Sınıf': '10',
    '11. Sınıf': '11',
    '12. Sınıf': '12',
    'YKS-TYT': 'tyt',
    'YKS-AYT': 'ayt'
}

def normalize_name(text):
    """Normalize topic name for comparison"""
    if not text:
        return ""
    # Lowercase
    text = text.lower()
    # Remove extra spaces
    text = re.sub(r'\s+', ' ', text).strip()
    # Remove special chars
    text = re.sub(r'[^\w\s:-]', '', text)
    return text

def extract_main_topic(text):
    """
    Extract main topic from full name
    "Matematik:Sayılar-Gerçek Sayılar..." → "sayılar gerçek sayılar"
    """
    if not text:
        return ""
    
    # Split by : (subject separator)
    if ':' in text:
        parts = text.split(':', 1)
        if len(parts) > 1:
            topic_part = parts[1]
        else:
            topic_part = text
    else:
        topic_part = text
    
    # Split by - (sub-topic separator) and take first 2-3 words
    if '-' in topic_part:
        words = topic_part.split('-')[0]
    else:
        words = topic_part
    
    return normalize_name(words)

def clean_value(val):
    """Clean value for SQL"""
    if val is None or val == '':
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0

def escape_sql(text):
    """Escape single quotes for SQL"""
    if not text:
        return ''
    return str(text).replace("'", "''")

def main():
    print("=" * 70)
    print("📊 SMART MATCHING SQL GENERATOR")
    print("=" * 70)
    
    # Load Excel
    print("\n📁 Loading Excel...")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb['Sayfa1']
        print(f"✅ Loaded: {ws.max_row - 3} rows")
    except Exception as e:
        print(f"❌ Failed: {e}")
        return
    
    # Generate SQL with multiple matching strategies
    print("\n📝 Generating SQL with smart matching...")
    
    sql_statements = []
    sql_statements.append("-- ============================================")
    sql_statements.append("-- UPDATE TOPICS WITH YEAR DATA")
    sql_statements.append("-- Smart name-based matching")
    sql_statements.append("-- ============================================")
    sql_statements.append("")
    sql_statements.append("BEGIN;")
    sql_statements.append("")
    
    # Create temp table for batch matching
    sql_statements.append("-- Create temp table for excel data")
    sql_statements.append("CREATE TEMP TABLE excel_topic_data (")
    sql_statements.append("    excel_id INT,")
    sql_statements.append("    subject_code TEXT,")
    sql_statements.append("    grade_level TEXT,")
    sql_statements.append("    topic_name TEXT,")
    sql_statements.append("    q_2018 DECIMAL(4,2),")
    sql_statements.append("    q_2019 DECIMAL(4,2),")
    sql_statements.append("    q_2020 DECIMAL(4,2),")
    sql_statements.append("    q_2021 DECIMAL(4,2),")
    sql_statements.append("    q_2022 DECIMAL(4,2),")
    sql_statements.append("    q_2023 DECIMAL(4,2),")
    sql_statements.append("    q_2024 DECIMAL(4,2),")
    sql_statements.append("    q_2025 DECIMAL(4,2)")
    sql_statements.append(");")
    sql_statements.append("")
    
    processed = 0
    skipped = 0
    
    for row_idx in range(4, ws.max_row + 1):
        # Read row (A=1, B=2, C=3, D=4, E=5...)
        konu_id = ws.cell(row_idx, 1).value  # A: KonuID
        sinif = ws.cell(row_idx, 2).value    # B: SINIF
        ders = ws.cell(row_idx, 3).value     # C: DERS
        ana_konu = ws.cell(row_idx, 4).value # D: ANA KONU
        
        if not all([konu_id, sinif, ders, ana_konu]):
            skipped += 1
            continue
        
        # Map subject
        subject_code = SUBJECT_MAPPING.get(ders)
        if not subject_code:
            skipped += 1
            continue
        
        # Map grade
        grade_mapped = GRADE_MAPPING.get(sinif, sinif)
        
        # Parse year data (E=5 to L=12)
        years = [2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025]
        year_values = {}
        
        for i, year in enumerate(years):
            col_idx = 5 + i  # E=5, F=6, G=7, H=8, I=9, J=10, K=11, L=12
            q_count = clean_value(ws.cell(row_idx, col_idx).value)
            year_values[f'q_{year}'] = q_count
        
        # Escape name
        safe_name = escape_sql(ana_konu)
        safe_grade = escape_sql(grade_mapped)
        
        # Insert into temp table
        sql = f"""INSERT INTO excel_topic_data VALUES (
    {konu_id},
    '{subject_code}',
    '{safe_grade}',
    '{safe_name}',
    {year_values['q_2018']}, {year_values['q_2019']}, {year_values['q_2020']}, {year_values['q_2021']},
    {year_values['q_2022']}, {year_values['q_2023']}, {year_values['q_2024']}, {year_values['q_2025']}
);
"""
        sql_statements.append(sql)
        processed += 1
        
        # Progress
        if row_idx % 200 == 0:
            print(f"   Processed {row_idx - 3}/{ws.max_row - 3}...")
    
    sql_statements.append("")
    sql_statements.append("-- ============================================")
    sql_statements.append("-- MATCHING STRATEGIES")
    sql_statements.append("-- ============================================")
    sql_statements.append("")
    
    # Strategy 1: Exact name + subject + grade match
    sql_statements.append("-- Strategy 1: Exact match (name + subject + grade)")
    sql_statements.append("""
UPDATE topics t
SET 
    q_2018 = e.q_2018,
    q_2019 = e.q_2019,
    q_2020 = e.q_2020,
    q_2021 = e.q_2021,
    q_2022 = e.q_2022,
    q_2023 = e.q_2023,
    q_2024 = e.q_2024,
    q_2025 = e.q_2025
FROM excel_topic_data e
JOIN subjects s ON s.code = e.subject_code
WHERE t.name_tr = e.topic_name
  AND t.subject_id = s.id
  AND (t.grade_level ILIKE '%' || e.grade_level || '%' OR e.grade_level ILIKE '%' || t.grade_level || '%');
""")
    sql_statements.append("")
    
    # Strategy 2: Partial name match (first 30 chars)
    sql_statements.append("-- Strategy 2: Partial match (first 30 characters)")
    sql_statements.append("""
UPDATE topics t
SET 
    q_2018 = COALESCE(t.q_2018, 0) + e.q_2018,
    q_2019 = COALESCE(t.q_2019, 0) + e.q_2019,
    q_2020 = COALESCE(t.q_2020, 0) + e.q_2020,
    q_2021 = COALESCE(t.q_2021, 0) + e.q_2021,
    q_2022 = COALESCE(t.q_2022, 0) + e.q_2022,
    q_2023 = COALESCE(t.q_2023, 0) + e.q_2023,
    q_2024 = COALESCE(t.q_2024, 0) + e.q_2024,
    q_2025 = COALESCE(t.q_2025, 0) + e.q_2025
FROM excel_topic_data e
JOIN subjects s ON s.code = e.subject_code
WHERE LEFT(t.name_tr, 30) = LEFT(e.topic_name, 30)
  AND t.subject_id = s.id
  AND t.q_2024 = 0  -- Only update if not already matched
  AND (t.grade_level ILIKE '%' || e.grade_level || '%' OR e.grade_level ILIKE '%' || t.grade_level || '%');
""")
    sql_statements.append("")
    
    # Strategy 3: Subject + grade only (for unmapped topics)
    sql_statements.append("-- Strategy 3: Display name match (if available)")
    sql_statements.append("""
UPDATE topics t
SET 
    q_2018 = COALESCE(t.q_2018, 0) + e.q_2018,
    q_2019 = COALESCE(t.q_2019, 0) + e.q_2019,
    q_2020 = COALESCE(t.q_2020, 0) + e.q_2020,
    q_2021 = COALESCE(t.q_2021, 0) + e.q_2021,
    q_2022 = COALESCE(t.q_2022, 0) + e.q_2022,
    q_2023 = COALESCE(t.q_2023, 0) + e.q_2023,
    q_2024 = COALESCE(t.q_2024, 0) + e.q_2024,
    q_2025 = COALESCE(t.q_2025, 0) + e.q_2025
FROM excel_topic_data e
JOIN subjects s ON s.code = e.subject_code
WHERE t.display_name = e.topic_name
  AND t.subject_id = s.id
  AND t.q_2024 = 0  -- Only update if not already matched
  AND (t.grade_level ILIKE '%' || e.grade_level || '%' OR e.grade_level ILIKE '%' || t.grade_level || '%');
""")
    sql_statements.append("")
    
    # Update exam_frequency
    sql_statements.append("-- ============================================")
    sql_statements.append("-- UPDATE EXAM FREQUENCY")
    sql_statements.append("-- ============================================")
    sql_statements.append("")
    sql_statements.append("-- Calculate frequency for all updated topics")
    sql_statements.append("SELECT update_all_exam_frequencies(5);")
    sql_statements.append("")
    
    # Cleanup
    sql_statements.append("-- Cleanup temp table")
    sql_statements.append("DROP TABLE excel_topic_data;")
    sql_statements.append("")
    sql_statements.append("COMMIT;")
    sql_statements.append("")
    
    # Verification
    sql_statements.append("-- ============================================")
    sql_statements.append("-- VERIFICATION")
    sql_statements.append("-- ============================================")
    sql_statements.append("")
    sql_statements.append("SELECT ")
    sql_statements.append("    COUNT(*) as total_topics,")
    sql_statements.append("    COUNT(*) FILTER (WHERE q_2024 > 0) as matched_2024,")
    sql_statements.append("    COUNT(*) FILTER (WHERE q_2023 > 0) as matched_2023,")
    sql_statements.append("    COUNT(*) FILTER (WHERE exam_frequency > 1.0) as with_frequency,")
    sql_statements.append("    ROUND(AVG(exam_frequency), 2) as avg_frequency")
    sql_statements.append("FROM topics;")
    sql_statements.append("")
    sql_statements.append("-- Top 10 high frequency topics")
    sql_statements.append("SELECT name_tr, grade_level, q_2023, q_2024, exam_frequency")
    sql_statements.append("FROM topics")
    sql_statements.append("WHERE exam_frequency > 1.0")
    sql_statements.append("ORDER BY exam_frequency DESC")
    sql_statements.append("LIMIT 10;")
    
    # Write to file
    with open(OUTPUT_SQL, 'w', encoding='utf-8') as f:
        f.write('\n'.join(sql_statements))
    
    print(f"\n✅ Generated SQL for {processed} topics")
    print(f"⚠️  Skipped {skipped} rows")
    print(f"\n📄 SQL file: {OUTPUT_SQL}")
    print(f"   Size: {os.path.getsize(OUTPUT_SQL) / 1024:.1f} KB")
    
    print("\n" + "=" * 70)
    print("✅ SMART MATCHING SQL READY!")
    print("=" * 70)
    print("\nFeatures:")
    print("  • 3 matching strategies (exact, partial, display_name)")
    print("  • Subject + Grade filtering")
    print("  • Automatic frequency calculation")
    print("  • Comprehensive verification")
    print("\nNext: Copy SQL to Supabase Editor and RUN!")

if __name__ == "__main__":
    main()
