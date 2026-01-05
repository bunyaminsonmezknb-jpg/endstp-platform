#!/usr/bin/env python3
"""
Excel to Topics UPDATE script
Updates existing topics with year-based question counts (q_2018-q_2025)
"""

import os
import sys
import openpyxl
from supabase import create_client

# ============================================
# CONFIGURATION
# ============================================
SUPABASE_URL = "https://runbsfxytxmtzweuaufr.supabase.co"

# Try environment variable first, then fallback to hardcoded
SUPABASE_KEY = os.getenv("eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InJ1bmJzZnh5dHhtdHp3ZXVhdWZyIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc2MzM4NzQ0OCwiZXhwIjoyMDc4OTYzNDQ4fQ.7KCeRiXLrnnH0ZXtE7ukk6XrrmGOoKE50gqZnD3SQrE")

if not SUPABASE_KEY:
    # Fallback: Use key directly (TEMPORARY - for testing only)
    SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InJ1bmJzZnh5dHhtdHp3ZXVhdWZyIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTcyOTExMzY3MiwiZXhwIjoyMDQ0Njg5NjcyfQ.IrKhU3mSfCTlbM97Auy3lQxQIl3QkxGc_jUtPTz5DXo"
    print("⚠️  Using hardcoded key (temporary)")

if not SUPABASE_KEY:
    print("❌ No API key available!")
    sys.exit(1)

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

# Excel file path (same directory as script)
import os
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, "derskonular.xlsx")

# ============================================
# SUBJECT MAPPING (Excel → Database)
# ============================================
SUBJECT_MAPPING = {
    'Matematik': 'MAT',
    'Fizik': 'FIZ',
    'Kimya': 'KIM',
    'Biyoloji': 'BIO',
    'Coğrafya': 'COG',
    'Tarih': 'TAR',
    'Türkçe': 'TUR',
    'Edebiyat': 'EDB',
    'Felsefe': 'FEL',
    'Din': 'DIN',
    'Mantık': 'MAN',
    'Psikoloji': 'PSI',
    'Sosyoloji': 'SOS'
}

# ============================================
# MATCHING STRATEGY
# ============================================
def normalize_text(text):
    """Normalize text for fuzzy matching"""
    if not text:
        return ""
    # Remove special chars, lowercase, strip spaces
    import re
    text = text.lower().strip()
    text = re.sub(r'[^\w\s]', '', text)  # Remove punctuation
    text = re.sub(r'\s+', ' ', text)  # Collapse spaces
    return text

def find_matching_topic(excel_row, db_topics):
    """
    Find matching topic in database using multiple strategies
    
    Strategy 1: Exact match on code (if exists)
    Strategy 2: Fuzzy match on name_tr + subject + grade
    Strategy 3: Partial name match
    """
    excel_name = normalize_text(excel_row['name'])
    excel_subject = excel_row['subject_code']
    excel_grade = excel_row['grade']
    
    # Strategy 1: Code match
    excel_code = f"{excel_subject}_{excel_row['id']}"
    for topic in db_topics:
        if topic['code'] == excel_code:
            return topic
    
    # Strategy 2: Fuzzy match (name + subject + grade)
    for topic in db_topics:
        db_name = normalize_text(topic['name_tr'])
        db_subject = topic.get('subject_id')
        db_grade = normalize_text(topic.get('grade_level', ''))
        
        # Check if names are similar (at least 70% match)
        if (excel_name in db_name or db_name in excel_name) and \
           db_subject == excel_row['subject_uuid'] and \
           excel_grade in db_grade:
            return topic
    
    # Strategy 3: Partial name match only
    for topic in db_topics:
        db_name = normalize_text(topic['name_tr'])
        if excel_name and db_name and (excel_name in db_name or db_name in excel_name):
            # Additional check: same subject
            if topic.get('subject_id') == excel_row['subject_uuid']:
                return topic
    
    return None

# ============================================
# MAIN FUNCTION
# ============================================
def main():
    print("=" * 70)
    print("📊 EXCEL TO TOPICS UPDATE SCRIPT")
    print("=" * 70)
    
    # Step 1: Load Excel
    print("\n📁 Loading Excel...")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb['Sayfa1']
        print(f"✅ Loaded: {ws.max_row - 3} rows")
    except Exception as e:
        print(f"❌ Failed: {e}")
        return
    
    # Step 2: Get subjects from database
    print("\n🔍 Fetching subjects...")
    try:
        subjects_result = supabase.table("subjects").select("id, code, name_tr").execute()
        subject_map = {s['code']: s['id'] for s in subjects_result.data}
        print(f"✅ Found {len(subject_map)} subjects")
    except Exception as e:
        print(f"❌ Failed: {e}")
        return
    
    # Step 3: Get ALL topics from database (for matching)
    print("\n🔍 Fetching topics...")
    try:
        topics_result = supabase.table("topics").select(
            "id, code, name_tr, subject_id, grade_level"
        ).execute()
        db_topics = topics_result.data
        print(f"✅ Found {len(db_topics)} topics in database")
    except Exception as e:
        print(f"❌ Failed: {e}")
        return
    
    # Step 4: Parse Excel
    print("\n📝 Parsing Excel and matching...")
    updates_to_apply = []
    unmatched = []
    
    for row_idx in range(4, ws.max_row + 1):
        # Read row
        konu_id = ws.cell(row_idx, 1).value
        sinif = ws.cell(row_idx, 2).value
        ders = ws.cell(row_idx, 3).value
        ana_konu = ws.cell(row_idx, 4).value
        
        if not all([konu_id, sinif, ders, ana_konu]):
            continue
        
        # Map subject
        subject_code = SUBJECT_MAPPING.get(ders)
        if not subject_code:
            continue
        
        subject_uuid = subject_map.get(subject_code)
        if not subject_uuid:
            continue
        
        # Prepare Excel row data
        excel_row = {
            'id': konu_id,
            'name': ana_konu,
            'subject_code': subject_code,
            'subject_uuid': subject_uuid,
            'grade': sinif
        }
        
        # Find matching topic
        matched_topic = find_matching_topic(excel_row, db_topics)
        
        if matched_topic:
            # Parse year data
            year_data = {}
            years = [2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025]
            for i, year in enumerate(years):
                col_idx = 5 + i
                q_count = ws.cell(row_idx, col_idx).value
                
                try:
                    q_count = float(q_count) if q_count is not None else 0.0
                except (ValueError, TypeError):
                    q_count = 0.0
                
                year_data[f'q_{year}'] = q_count
            
            updates_to_apply.append({
                'topic_id': matched_topic['id'],
                'topic_name': matched_topic['name_tr'],
                'year_data': year_data
            })
        else:
            unmatched.append({
                'excel_name': ana_konu,
                'subject': ders,
                'grade': sinif
            })
        
        # Progress
        if row_idx % 100 == 0:
            print(f"   Processed {row_idx - 3}/{ws.max_row - 3}...")
    
    print(f"\n✅ Matched: {len(updates_to_apply)} topics")
    print(f"⚠️  Unmatched: {len(unmatched)} topics")
    
    if unmatched:
        print("\n⚠️  First 5 unmatched topics:")
        for item in unmatched[:5]:
            print(f"   • {item['excel_name'][:50]} ({item['subject']}, {item['grade']})")
    
    # Step 5: Apply updates (batch)
    print(f"\n💾 Updating {len(updates_to_apply)} topics...")
    BATCH_SIZE = 50
    updated_count = 0
    failed_count = 0
    
    for i in range(0, len(updates_to_apply), BATCH_SIZE):
        batch = updates_to_apply[i:i+BATCH_SIZE]
        
        for update in batch:
            try:
                supabase.table("topics").update(
                    update['year_data']
                ).eq('id', update['topic_id']).execute()
                
                updated_count += 1
            except Exception as e:
                print(f"   ❌ Failed: {update['topic_name'][:30]} - {e}")
                failed_count += 1
        
        print(f"   Updated {updated_count}/{len(updates_to_apply)}...")
    
    print(f"\n✅ Successfully updated: {updated_count}")
    if failed_count > 0:
        print(f"❌ Failed: {failed_count}")
    
    # Step 6: Update exam_frequency for all topics
    print("\n🔄 Updating exam_frequency (this may take 30 seconds)...")
    try:
        result = supabase.rpc('update_all_exam_frequencies', {'p_years_back': 5}).execute()
        
        if result.data:
            stats = result.data[0]
            print(f"✅ Updated {stats['updated_count']} topics")
            print(f"   Avg frequency: {stats['avg_frequency']}")
            print(f"   Max frequency: {stats['max_frequency']}")
            print(f"   Min frequency: {stats['min_frequency']}")
    except Exception as e:
        print(f"⚠️  Auto-update failed: {e}")
        print("   Run manually: SELECT update_all_exam_frequencies(5);")
    
    # Step 7: Verification
    print("\n🔍 Verification:")
    print("-" * 70)
    
    # Count topics with data
    verify_query = """
    SELECT 
        COUNT(*) as total,
        COUNT(*) FILTER (WHERE q_2024 > 0) as with_2024,
        COUNT(*) FILTER (WHERE q_2023 > 0) as with_2023,
        COUNT(*) FILTER (WHERE exam_frequency IS NOT NULL) as with_freq
    FROM topics;
    """
    
    try:
        verify_result = supabase.rpc('query', {'query': verify_query}).execute()
        if verify_result.data:
            v = verify_result.data[0]
            print(f"  Total topics: {v.get('total', 'N/A')}")
            print(f"  Topics with 2024 data: {v.get('with_2024', 'N/A')}")
            print(f"  Topics with 2023 data: {v.get('with_2023', 'N/A')}")
            print(f"  Topics with frequency: {v.get('with_freq', 'N/A')}")
    except:
        # Fallback
        count_result = supabase.table("topics").select("*", count='exact').execute()
        print(f"  Total topics: {count_result.count}")
    
    # Sample topics
    print("\n📝 Sample topics with frequency:")
    print("-" * 70)
    sample = supabase.table("topics").select(
        "name_tr, q_2023, q_2024, exam_frequency"
    ).not_.is_('exam_frequency', 'null').limit(5).execute()
    
    if sample.data:
        for topic in sample.data:
            print(f"  {topic['name_tr'][:40]:40} | 2023:{topic['q_2023']:4.1f} 2024:{topic['q_2024']:4.1f} | Freq:{topic['exam_frequency']:4.2f}")
    
    print("\n" + "=" * 70)
    print("✅ UPDATE COMPLETE!")
    print("=" * 70)
    print("\n🎯 Next: Test Priority Engine v2 with frequency scores!")

if __name__ == "__main__":
    main()